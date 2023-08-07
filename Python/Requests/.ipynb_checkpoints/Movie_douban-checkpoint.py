import datetime
from openpyxl import Workbook
import random
import re
import time
import requests

wb = Workbook()
ws = wb.active

ws['A1'] = '电影名称'
ws['B1'] = '豆瓣评分'
ws['C1'] = '日期'

filename = "./Data/movie_douban_"+str(time.strftime('%Y-%m-%d_%H%M%S')) + ".xlsx"

for page in range(1, 11):
    resp = requests.get(
        url=f'https://movie.douban.com/top250?start={(page - 1) * 25}',
        # 如果不设置HTTP请求头中的User-Agent，豆瓣会检测出不是浏览器而阻止我们的请求。
        headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36'}
    )
    # 通过正则表达式获取class属性为title且标签体不以&开头的span标签并用捕获组提取标签内容
    pattern1 = re.compile(r'<span class="title">([^&]*?)</span>')
    titles = pattern1.findall(resp.text)
    # 通过正则表达式获取class属性为rating_num的span标签并用捕获组提取标签内容
    pattern2 = re.compile(r'<span class="rating_num".*?>(.*?)</span>')
    ranks = pattern2.findall(resp.text)
    # 使用zip压缩两个列表，循环遍历所有的电影标题和评分
    for title, rank in zip(titles, ranks):
        print(title, rank)
        ws.append([title, rank, datetime.datetime.now().strftime('%Y-%m-%d')])
    # 随机休眠1-5秒，避免爬取页面过于频繁
    time.sleep(random.random() * 4 + 1)


wb.save(filename)